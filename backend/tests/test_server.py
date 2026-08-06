import pytest
import io
import openpyxl
from unittest.mock import patch
from fastapi.testclient import TestClient

@pytest.fixture
def test_storage(tmp_path):
    from storage import SQLiteStorage
    test_db = tmp_path / "test_server_store.db"
    return SQLiteStorage(str(test_db))

@pytest.fixture
def client(test_storage):
    import server
    
    # Patch storage.create_storage so lifespan registers test_storage
    async def mock_create():
        return test_storage, "test-json-file"
        
    with patch("storage.create_storage", side_effect=mock_create):
        with TestClient(server.app) as c:
            yield c


def test_health_check(client):
    response = client.get("/api/")
    assert response.status_code == 200
    data = response.json()
    assert data["app"] == "Batua"
    assert data["status"] == "live"
    assert data["storage"] == "test-json-file"

def test_parse_nl_route(client):
    response = client.post("/api/parse-nl", json={"text": "zomato 450 yesterday upi"})
    assert response.status_code == 200
    data = response.json()
    assert data["description"] == "Zomato"
    assert data["amount"] == -450.0
    assert data["category"] == "Food Delivery"
    assert data["payment_method"] == "UPI"

    # Empty payload error
    response = client.post("/api/parse-nl", json={"text": "   "})
    assert response.status_code == 400


def test_parse_voice_route(client):
    response = client.post(
        "/api/parse-nl/voice",
        json={
            "text": "aaj maine 11 bje kurkure ka packet liya 10 wala fer 2 bje din k gol gappe khaye 20 k"
        },
    )
    assert response.status_code == 200
    items = response.json()["items"]
    assert len(items) == 2
    assert items[0]["description"] == "Kurkure Packet"
    assert items[0]["amount"] == -10.0
    assert items[0]["category"] == "Snacks"
    # Time is stripped but not stored — only the date is kept.
    assert items[0].get("notes", "") == ""
    assert items[1]["description"] == "Gol Gappe"
    assert items[1]["amount"] == -20.0
    assert items[1]["category"] == "Snacks"
    assert items[1].get("notes", "") == ""


def test_parse_voice_route_quantity_and_multiprice(client):
    response = client.post(
        "/api/parse-nl/voice",
        json={"text": "aaj maine lays k 2 packet liye ek 10 ka ek 20 ka aur chai 10"},
    )
    assert response.status_code == 200
    items = response.json()["items"]
    assert len(items) == 2
    lays = items[0]
    assert lays["quantity"] == 2
    assert lays["amount"] == -30.0
    assert lays["category"] == "Snacks"
    assert items[1]["amount"] == -10.0


def test_transcribe_status_route(client):
    response = client.get("/api/transcribe/status")
    assert response.status_code == 200
    data = response.json()
    assert "available" in data
    assert "model" in data


def test_transcribe_route_success(client):
    # Mock the whisper engine so the test needs no audio model / real decoding.
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.transcribe_file", return_value="chai 10 aur samosa 15"):
        response = client.post(
            "/api/transcribe",
            files={"file": ("voice.webm", b"fake-audio-bytes", "audio/webm")},
        )
    assert response.status_code == 200
    data = response.json()
    assert data["text"] == "chai 10 aur samosa 15"
    # The transcript is run through the same voice parser -> two items.
    assert len(data["items"]) == 2
    assert data["items"][0]["amount"] == -10.0
    assert data["items"][1]["amount"] == -15.0


def test_transcribe_route_unavailable(client):
    with patch("transcribe.is_enabled", return_value=False):
        response = client.post(
            "/api/transcribe",
            files={"file": ("voice.webm", b"fake-audio-bytes", "audio/webm")},
        )
    assert response.status_code == 503


def test_transcribe_route_empty_audio(client):
    with patch("transcribe.is_enabled", return_value=True):
        response = client.post(
            "/api/transcribe",
            files={"file": ("voice.webm", b"", "audio/webm")},
        )
    assert response.status_code == 400


def test_transcribe_status_lists_models(client):
    response = client.get("/api/transcribe/status")
    data = response.json()
    # The Settings mic-test panel needs the choices + which are loaded.
    assert isinstance(data["models"], list) and data["model"] in data["models"]
    assert isinstance(data["loaded"], list)


def test_transcribe_set_model(client):
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.set_active_model", return_value="base") as set_model:
        response = client.post("/api/transcribe/model", json={"model": "base"})
    assert response.status_code == 200
    assert response.json()["model"] == "base"
    set_model.assert_called_once_with("base")


def test_transcribe_set_model_rejects_unknown(client):
    with patch("transcribe.is_enabled", return_value=True):
        response = client.post("/api/transcribe/model", json={"model": "bogus"})
    assert response.status_code == 400


def test_transcribe_test_route_returns_details(client):
    # Mic test returns raw text + detection metadata, NOT parsed transactions.
    details = {
        "text": "chai 10 aur samosa 15",
        "model": "small",
        "language": "hi",
        "language_probability": 0.98,
        "duration_ms": 1200,
    }
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.transcribe_details", return_value=details):
        response = client.post(
            "/api/transcribe/test",
            files={"file": ("voice.webm", b"fake-audio-bytes", "audio/webm")},
            data={"model": "small"},
        )
    assert response.status_code == 200
    data = response.json()
    assert data == details
    assert "items" not in data  # unlike /transcribe, no parsing here


def test_transcribe_test_route_unavailable(client):
    with patch("transcribe.is_enabled", return_value=False):
        response = client.post(
            "/api/transcribe/test",
            files={"file": ("voice.webm", b"fake-audio-bytes", "audio/webm")},
        )
    assert response.status_code == 503


def test_transcribe_warm_success(client):
    """Warming a known model that loads fine returns 200 with model + loaded list."""
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.model_name", return_value="small"), \
         patch("transcribe.available_models", return_value=["tiny", "base", "small", "medium", "large-v3"]), \
         patch("transcribe.warm_model", return_value=True) as warm, \
         patch("transcribe.loaded_models", return_value=["small"]):
        response = client.post("/api/transcribe/warm", json={"model": "small"})
    assert response.status_code == 200
    data = response.json()
    assert data["model"] == "small"
    assert "loaded" in data
    warm.assert_called_once_with("small")


def test_transcribe_warm_default_model(client):
    """Omitting the model payload warms the currently active model."""
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.model_name", return_value="base"), \
         patch("transcribe.available_models", return_value=["tiny", "base", "small", "medium", "large-v3"]), \
         patch("transcribe.warm_model", return_value=True) as warm, \
         patch("transcribe.loaded_models", return_value=["base"]):
        # Send an empty JSON body — the route should fall back to model_name().
        response = client.post("/api/transcribe/warm")
    assert response.status_code == 200
    assert response.json()["model"] == "base"
    warm.assert_called_once_with("base")


def test_transcribe_warm_rejects_unknown(client):
    """An unknown model size returns 400."""
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.model_name", return_value="small"), \
         patch("transcribe.available_models", return_value=["tiny", "base", "small", "medium", "large-v3"]):
        response = client.post("/api/transcribe/warm", json={"model": "bogus"})
    assert response.status_code == 400


def test_transcribe_warm_load_failure(client):
    """If warm_model returns False (load failed), the route returns 500."""
    with patch("transcribe.is_enabled", return_value=True), \
         patch("transcribe.model_name", return_value="small"), \
         patch("transcribe.available_models", return_value=["tiny", "base", "small", "medium", "large-v3"]), \
         patch("transcribe.warm_model", return_value=False):
        response = client.post("/api/transcribe/warm", json={"model": "small"})
    assert response.status_code == 500


def test_transcribe_warm_unavailable(client):
    """If offline transcription is disabled, the route returns 503."""
    with patch("transcribe.is_enabled", return_value=False):
        response = client.post("/api/transcribe/warm", json={"model": "small"})
    assert response.status_code == 503


def test_transaction_crud(client):
    # 1. Create a transaction
    payload = {
        "date": "2026-06-19",
        "description": "Zomato lunch",
        "amount": -450.0,
        "category": "Food & Dining",
        "payment_method": "UPI",
        "notes": "Spicy paneer biryani"
    }
    response = client.post("/api/transactions", json=payload)
    assert response.status_code == 200
    txn = response.json()
    assert txn["description"] == "Zomato lunch"
    assert txn["amount"] == -450.0
    assert txn["txn_type"] == "debit"
    assert "id" in txn

    txn_id = txn["id"]

    # 2. Get list of transactions
    response = client.get("/api/transactions")
    assert response.status_code == 200
    list_data = response.json()
    assert list_data["total"] == 1
    assert list_data["items"][0]["id"] == txn_id

    # 3. Update transaction
    response = client.put(f"/api/transactions/{txn_id}", json={"amount": -500.0, "notes": "Price increased"})
    assert response.status_code == 200
    updated_txn = response.json()
    assert updated_txn["amount"] == -500.0
    assert updated_txn["notes"] == "Price increased"
    assert updated_txn["txn_type"] == "debit"

    # Try non-existent transaction update
    response = client.put("/api/transactions/non-existent-id", json={"amount": 100})
    assert response.status_code == 404

    # 4. Get transaction list with filter
    response = client.get("/api/transactions", params={"search": "Zomato", "category": "Food & Dining"})
    assert len(response.json()["items"]) == 1

    # Search with no hit
    response = client.get("/api/transactions", params={"search": "Uber"})
    assert len(response.json()["items"]) == 0

    # 5. Delete transaction
    response = client.delete(f"/api/transactions/{txn_id}")
    assert response.status_code == 200
    assert response.json() == {"deleted": 1}

    # Delete non-existent
    response = client.delete("/api/transactions/non-existent-id")
    assert response.status_code == 404


def test_transaction_create_clamps_future_date(client):
    """A future-dated transaction is stored as today, never in the future."""
    response = client.post(
        "/api/transactions",
        json={"date": "2099-01-01", "description": "Future thing", "amount": -100.0, "category": "Other"},
    )
    assert response.status_code == 200
    txn = response.json()
    from datetime import datetime
    assert txn["date"] <= datetime.now().strftime("%Y-%m-%d")
    assert txn["date"] != "2099-01-01"


def test_transaction_update_clamps_future_date(client):
    """Editing a transaction to a future date is clamped to today."""
    created = client.post(
        "/api/transactions",
        json={"date": "2026-06-10", "description": "T", "amount": -50.0},
    ).json()
    r = client.put(f"/api/transactions/{created['id']}", json={"date": "2099-12-31"})
    assert r.status_code == 200
    from datetime import datetime
    assert r.json()["date"] <= datetime.now().strftime("%Y-%m-%d")


def test_gemini_key_setting_route(client):
    # "ai" here is the project's own backend/ai.py Gemini wrapper, not the
    # google.generativeai SDK. Patches use string paths so no import is needed.
    # Status reflects whether a key is configured.
    with patch("ai.is_enabled", return_value=False):
        r = client.get("/api/settings/gemini-key")
        assert r.status_code == 200
        assert r.json() == {"configured": False}

    # A valid key (per validate_key) is persisted with the trimmed value.
    with patch("ai.validate_key", return_value=(True, "ok", "Key valid")), patch(
        "ai.set_api_key"
    ) as set_key:
        r = client.put("/api/settings/gemini-key", json={"api_key": "  AIza-valid  "})
        assert r.status_code == 200
        assert r.json()["updated"] is True
        assert r.json()["configured"] is True
        set_key.assert_called_once_with("AIza-valid")

    # Empty/whitespace keys are rejected.
    r = client.put("/api/settings/gemini-key", json={"api_key": "   "})
    assert r.status_code == 400

    # A key the Gemini API rejects must NOT be persisted and returns 400 with a
    # distinct reason — a bad key can never silently replace a good one.
    with patch("ai.validate_key", return_value=(False, "invalid_key", "Key rejected")), patch(
        "ai.set_api_key"
    ) as set_key:
        r = client.put("/api/settings/gemini-key", json={"api_key": "AIza-fake"})
        assert r.status_code == 400
        assert r.json()["detail"]["reason"] == "invalid_key"
        set_key.assert_not_called()

    # A network/transient failure is surfaced separately from auth failure.
    with patch("ai.validate_key", return_value=(False, "network_error", "Timeout")), patch(
        "ai.set_api_key"
    ) as set_key:
        r = client.put("/api/settings/gemini-key", json={"api_key": "AIza-temp"})
        assert r.status_code == 502
        assert r.json()["detail"]["reason"] == "network_error"
        set_key.assert_not_called()

    # Test endpoint reports the same distinct reasons.
    with patch("ai.validate_key", return_value=(False, "invalid_key", "Key rejected")):
        r = client.post("/api/settings/gemini-key/test")
        assert r.status_code == 200
        assert r.json() == {"valid": False, "reason": "invalid_key", "message": "Key rejected"}


def test_validate_key_retries_transient_and_distinguishes_reasons(client):
    """A valid key must not be rejected by a single transient failure, and
    auth/network/quota reasons stay distinct."""
    import ai as ai_mod

    class _Resp:
        def __init__(self, code, json_body=None, ct="application/json"):
            self.status_code = code
            self._json = json_body or {}
            self.headers = {"content-type": ct}

        def json(self):
            return self._json

    # 500 → 200: retried once, succeeds.
    with patch("ai.requests.get", side_effect=[_Resp(500), _Resp(200)]) as get:
        ok, reason, _ = ai_mod.validate_key("AIza-valid")
        assert ok is True and reason == "ok"
        assert get.call_count == 2

    # Connection exception → 200: retried once, succeeds.
    with patch("ai.requests.get", side_effect=[ConnectionError("boom"), _Resp(200)]) as get:
        ok, reason, _ = ai_mod.validate_key("AIza-valid")
        assert ok is True and reason == "ok"
        assert get.call_count == 2

    # 400 auth failure is authoritative — no retry.
    with patch("ai.requests.get", side_effect=[_Resp(400, {"error": {"message": "API key not valid"}})]) as get:
        ok, reason, msg = ai_mod.validate_key("AIza-bad")
        assert ok is False and reason == "invalid_key"
        assert get.call_count == 1
        assert "API key not valid" in msg

    # 429 → quota, distinct from a network error.
    with patch("ai.requests.get", side_effect=[_Resp(429), _Resp(429)]) as get:
        ok, reason, _ = ai_mod.validate_key("AIza-valid")
        assert ok is False and reason == "quota"
        assert get.call_count == 2

    # Quota reason maps to HTTP 429 on the settings route.
    with patch("ai.validate_key", return_value=(False, "quota", "Rate limited")):
        r = client.put("/api/settings/gemini-key", json={"api_key": "AIza-valid"})
        assert r.status_code == 429
        assert r.json()["detail"]["reason"] == "quota"

    # 404 → the key is valid but the model isn't available — its own reason,
    # with a friendly message that never leaks a raw exception.
    with patch("ai.requests.get", side_effect=[_Resp(404)]) as get:
        ok, reason, msg = ai_mod.validate_key("AIza-valid")
        assert ok is False and reason == "model_unavailable"
        assert get.call_count == 1  # authoritative, not retried
        assert "GEMINI_MODEL" in msg

    # Network failure → friendly message (no raw exception string); retried once.
    with patch("ai.requests.get", side_effect=[ConnectionError("some raw detail"), ConnectionError("some raw detail")]) as get:
        ok, reason, msg = ai_mod.validate_key("AIza-valid")
        assert ok is False and reason == "network_error"
        assert get.call_count == 2
        assert "some raw detail" not in msg


def test_transaction_price_derivation(client):
    """Test that POST without price derives price and PUT changing quantity recomputes price."""
    # 1. POST without price - should derive price = |amount|/quantity
    payload = {
        "date": "2026-06-19",
        "description": "Lays chips",
        "amount": -90.0,
        "quantity": 3,
        "category": "Snacks"
    }
    response = client.post("/api/transactions", json=payload)
    assert response.status_code == 200
    txn = response.json()
    assert txn["price"] == 30.0  # 90/3 = 30
    
    txn_id = txn["id"]
    
    # 2. PUT changing quantity without explicit price - should recompute price
    response = client.put(f"/api/transactions/{txn_id}", json={"quantity": 2})
    assert response.status_code == 200
    updated = response.json()
    assert updated["quantity"] == 2
    assert updated["price"] == 45.0  # 90/2 = 45
    
    # 3. PUT with explicit price - should use provided price
    response = client.put(f"/api/transactions/{txn_id}", json={"price": 40.0})
    assert response.status_code == 200
    updated = response.json()
    assert updated["price"] == 40.0

def test_bulk_delete(client):
    # Insert two transactions
    t1 = client.post("/api/transactions", json={"date": "2026-06-19", "description": "T1", "amount": -10}).json()
    t2 = client.post("/api/transactions", json={"date": "2026-06-19", "description": "T2", "amount": -20}).json()
    
    # Check total is 2
    assert client.get("/api/transactions").json()["total"] == 2
    
    # Bulk delete them
    response = client.post("/api/transactions/bulk-delete", json={"ids": [t1["id"], t2["id"]]})
    assert response.status_code == 200
    assert response.json()["deleted"] == 2

    # Check total is 0
    assert client.get("/api/transactions").json()["total"] == 0

def test_wipe_transactions(client):
    client.post("/api/transactions", json={"date": "2026-06-19", "description": "T1", "amount": -10})
    client.post("/api/transactions", json={"date": "2026-06-19", "description": "T2", "amount": -20})
    
    response = client.delete("/api/transactions")
    assert response.status_code == 200
    assert response.json()["deleted"] == 2
    assert client.get("/api/transactions").json()["total"] == 0

def test_budgets_crud(client):
    # Create budget
    response = client.post("/api/budgets", json={"category": "Food & Dining", "limit": 5000.0})
    assert response.status_code == 200
    budget = response.json()
    assert budget["category"] == "Food & Dining"
    assert budget["limit"] == 5000.0
    budget_id = budget["id"]

    # List budgets
    response = client.get("/api/budgets")
    assert response.status_code == 200
    assert len(response.json()["budgets"]) == 1
    assert response.json()["budgets"][0]["id"] == budget_id

    # Update budget limit (upsert)
    response = client.post("/api/budgets", json={"category": "Food & Dining", "limit": 6000.0})
    assert response.status_code == 200
    # List again to verify limit updated
    response = client.get("/api/budgets")
    assert response.json()["budgets"][0]["limit"] == 6000.0

    # Delete budget
    response = client.delete(f"/api/budgets/{budget_id}")
    assert response.status_code == 200
    assert response.json() == {"deleted": 1}

    # Delete non-existent
    response = client.delete("/api/budgets/non-existent-id")
    assert response.status_code == 404

def test_dashboard_metrics_and_analytics(client):
    # Insert test data:
    # Current month: 2026-06
    # Previous month: 2026-05
    client.post("/api/transactions", json={"date": "2026-06-10", "description": "Salary", "amount": 10000.0, "category": "Income"})
    client.post("/api/transactions", json={"date": "2026-06-11", "description": "Rent", "amount": -2000.0, "category": "Housing/Rent"})
    client.post("/api/transactions", json={"date": "2026-06-12", "description": "Swiggy", "amount": -500.0, "category": "Food & Dining"})
    
    # Prev Month
    client.post("/api/transactions", json={"date": "2026-05-10", "description": "Salary", "amount": 8000.0, "category": "Income"})
    client.post("/api/transactions", json={"date": "2026-05-12", "description": "Swiggy", "amount": -400.0, "category": "Food & Dining"})

    # Dashboard Metrics
    response = client.get("/api/dashboard/metrics")
    assert response.status_code == 200
    metrics = response.json()
    assert metrics["income"] == 10000.0
    assert metrics["expense"] == 2500.0
    assert metrics["net"] == 7500.0
    assert metrics["savings_rate"] == 75.0
    assert metrics["current_month"] == "2026-06"
    # Percentage changes:
    # Income: 10000 vs 8000 (+25%)
    # Expense: 2500 vs 400 (+525%)
    assert metrics["income_change"] == 25.0
    assert metrics["expense_change"] == 525.0

    # Analytics Timeline
    response = client.get("/api/analytics/timeline")
    assert response.status_code == 200
    series = response.json()["series"]
    assert len(series) == 2
    assert series[0]["month"] == "2026-05"
    assert series[0]["income"] == 8000.0
    assert series[1]["month"] == "2026-06"
    assert series[1]["income"] == 10000.0

    # Category Breakdown
    response = client.get("/api/analytics/category-breakdown", params={"month": "2026-06"})
    assert response.status_code == 200
    cats = response.json()["data"]
    # Should have Housing/Rent and Food & Dining
    assert len(cats) == 2
    assert cats[0]["category"] == "Housing/Rent"
    assert cats[0]["amount"] == 2000.0

    # Top Merchants
    response = client.get("/api/analytics/top-merchants")
    assert response.status_code == 200
    merchants = response.json()["data"]
    assert merchants[0]["merchant"] == "Rent"
    assert merchants[0]["amount"] == 2000.0

    # Heatmap
    response = client.get("/api/analytics/heatmap")
    assert response.status_code == 200
    assert "days" in response.json()
    assert "max" in response.json()

    # Payment Method
    response = client.get("/api/analytics/payment-method")
    assert response.status_code == 200
    # None of the txns had payment methods, should aggregate under "Unknown" or empty
    assert len(response.json()["data"]) >= 1

    # Treemap
    response = client.get("/api/analytics/treemap")
    assert response.status_code == 200
    assert len(response.json()["data"]) == 2  # Housing/Rent & Food & Dining

    # Categories list
    response = client.get("/api/categories")
    assert response.status_code == 200
    assert "Income" in response.json()["categories"]

    # Budget Status
    client.post("/api/budgets", json={"category": "Food & Dining", "limit": 1000.0})
    response = client.get("/api/budgets/status", params={"month": "2026-06"})
    assert response.status_code == 200
    status_rows = response.json()["rows"]
    assert len(status_rows) == 1
    assert status_rows[0]["category"] == "Food & Dining"
    assert status_rows[0]["spent"] == 500.0
    assert status_rows[0]["pct"] == 50.0
    assert status_rows[0]["status"] == "ok"

    # Recurring Expenses
    response = client.get("/api/recurring")
    assert response.status_code == 200
    # No merchant appears in >= 3 months yet, so empty list
    assert len(response.json()["recurring"]) == 0

    # Insights
    response = client.get("/api/insights")
    assert response.status_code == 200
    assert "insights" in response.json()

def test_export_endpoints(client):
    client.post("/api/transactions", json={"date": "2026-06-10", "description": "Salary", "amount": 10000.0, "category": "Income"})
    
    # CSV Export
    response = client.get("/api/export/csv")
    assert response.status_code == 200
    assert response.headers["Content-Disposition"] == "attachment; filename=batua_transactions.csv"
    assert "Salary" in response.text
    
    # Excel Export (stacked expense-sheet format, now with income + Debit/Credit)
    response = client.get("/api/export/excel")
    assert response.status_code == 200
    assert response.headers["Content-Disposition"] == "attachment; filename=batua_expenditure.xlsx"
    assert len(response.content) > 0
    # Income alone still produces a workbook containing one (credit) block.
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["Expenses"]
    ws = wb["Expenses"]
    a_vals = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Expense Table : 06/2026" in a_vals
    assert "Salary" in [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]

def test_export_month_and_range(client):
    # Two expenses in May, two + income in June.
    for d, desc, amt in [
        ("2026-05-03", "Petrol", -1150.0),
        ("2026-05-11", "Zomato Dinner", -520.0),
        ("2026-06-02", "Netflix", -799.0),
        ("2026-06-14", "Swiggy Lunch", -380.0),
        ("2026-06-10", "Salary", 10000.0),
    ]:
        client.post("/api/transactions", json={"date": d, "description": desc, "amount": amt, "category": "Other"})

    # Months endpoint — income-only months never appear; newest first.
    response = client.get("/api/export/months")
    assert response.status_code == 200
    assert response.json()["months"] == ["2026-06", "2026-05"]

    # Single-month export → one block: that month's expenses AND income, and the
    # month's TOTAL (column E, Total Amount) equals the debit subtotal only —
    # income is not netted out.
    response = client.get("/api/export/excel", params={"month": "2026-06"})
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    ws = wb["Expenses"]
    a1 = ws["A1"].value
    assert a1 == "Expense Table : 06/2026"
    descs = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "Netflix" in descs and "Swiggy Lunch" in descs and "Salary" in descs
    assert "Petrol" not in descs
    # Salary is a Credit row; expenses are Debit rows.
    types = [ws.cell(row=r, column=3).value for r in range(1, ws.max_row + 1)]
    assert "Credit" in types and types.count("Debit") == 2
    total = ws.cell(row=ws.max_row, column=5).value
    assert total == 799.0 + 380.0
    # Quantity column is gone; header band is the navy "navigation" style.
    headers = [ws.cell(row=2, column=c_).value for c_ in range(1, ws.max_column + 1)]
    assert "Quantity" not in headers
    assert ws.cell(row=2, column=1).fill.fgColor.rgb == "FF2C3E50"
    # Dates are real dd/mm/yyyy cells, not mm-dd-yy text.
    assert ws.cell(row=3, column=6).number_format == "dd/mm/yyyy"
    # Zebra striping — second data row has the light-gray fill.
    assert ws.cell(row=4, column=5).fill.fgColor.rgb == "FFF7F7F5"

    # Custom date range export → only rows inside the inclusive range.
    response = client.get("/api/export/excel", params={"from": "2026-05-01", "to": "2026-05-31"})
    assert response.status_code == 200
    ws = openpyxl.load_workbook(io.BytesIO(response.content))["Expenses"]
    descs = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "Petrol" in descs and "Zomato Dinner" in descs
    assert "Netflix" not in descs
    # May + June together (no params) → two blocks with a YEAR marker between.
    response = client.get("/api/export/excel")
    ws = openpyxl.load_workbook(io.BytesIO(response.content))["Expenses"]
    a_vals = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Expense Table : 05/2026" in a_vals
    assert "Expense Table : 06/2026" in a_vals

def test_excel_upload_endpoints(client):
    # 1. Preview
    csv_data = "Date,Particulars,Amount\n12/06/2026,Zomato Lunch,-450.00\n"
    file_payload = {"file": ("statement.csv", io.BytesIO(csv_data.encode("utf-8")), "text/csv")}
    
    response = client.post("/api/upload-excel/preview", files=file_payload)
    assert response.status_code == 200
    data = response.json()
    assert data["format"] == "tabular"
    assert data["mapping"]["description"] == "Particulars"

    # 2. Upload
    file_payload = {"file": ("statement.csv", io.BytesIO(csv_data.encode("utf-8")), "text/csv")}
    response = client.post("/api/upload-excel", files=file_payload, params={"replace": True, "use_ai": False})
    assert response.status_code == 200
    assert response.json()["inserted"] == 1
    assert response.json()["replaced"] is True

    # Check that transaction was added
    txns = client.get("/api/transactions").json()
    assert txns["total"] == 1
    assert txns["items"][0]["description"] == "Zomato Lunch"


# --------------------------------------------------------------------------- #
# People Ledger
# --------------------------------------------------------------------------- #

def test_people_list_empty(client):
    response = client.get("/api/people/")
    assert response.status_code == 200
    assert response.json() == {"entries": []}


def test_people_create_and_list(client):
    response = client.post(
        "/api/people/",
        json={
            "person_name": "Rahul",
            "direction": "gave",
            "amount": 500,
            "reason": "lunch",
            "date": "2026-07-10",
        },
    )
    assert response.status_code == 200
    created = response.json()
    assert created["person_name"] == "Rahul"
    assert created["direction"] == "gave"
    assert created["amount"] == 500.0
    assert created["settled"] is False
    assert created["id"].startswith("pe_")
    assert created["created_at"]

    response = client.get("/api/people/")
    assert response.status_code == 200
    entries = response.json()["entries"]
    assert len(entries) == 1
    assert entries[0]["person_name"] == "Rahul"


def test_people_rejects_bad_inputs(client):
    # Empty name
    r = client.post("/api/people/", json={"person_name": "  ", "direction": "gave", "amount": 10, "date": "2026-07-10"})
    assert r.status_code == 400
    # Bad direction
    r = client.post("/api/people/", json={"person_name": "X", "direction": "sideways", "amount": 10, "date": "2026-07-10"})
    assert r.status_code == 400
    # Non-positive amount
    r = client.post("/api/people/", json={"person_name": "X", "direction": "gave", "amount": 0, "date": "2026-07-10"})
    assert r.status_code == 400
    r = client.post("/api/people/", json={"person_name": "X", "direction": "gave", "amount": -50, "date": "2026-07-10"})
    assert r.status_code == 400
    # Bad date
    r = client.post("/api/people/", json={"person_name": "X", "direction": "gave", "amount": 10, "date": "yesterday"})
    assert r.status_code == 400


def test_people_update_partial(client):
    create = client.post(
        "/api/people/",
        json={"person_name": "Mom", "direction": "took", "amount": 1000, "date": "2026-07-10", "reason": "rent share"},
    ).json()
    eid = create["id"]

    # Mark settled (single-field patch)
    r = client.put(f"/api/people/{eid}", json={"settled": True})
    assert r.status_code == 200
    assert r.json()["settled"] is True
    # Other fields unchanged
    assert r.json()["person_name"] == "Mom"
    assert r.json()["amount"] == 1000.0

    # Edit amount + reason
    r = client.put(f"/api/people/{eid}", json={"amount": 1500, "reason": "rent + groceries"})
    assert r.status_code == 200
    assert r.json()["amount"] == 1500.0
    assert r.json()["reason"] == "rent + groceries"
    assert r.json()["settled"] is True  # still settled from earlier patch

    # Unknown id -> 404
    r = client.put("/api/people/pe_doesnotexist", json={"settled": True})
    assert r.status_code == 404


def test_people_delete(client):
    create = client.post(
        "/api/people/",
        json={"person_name": "A", "direction": "gave", "amount": 100, "date": "2026-07-10"},
    ).json()
    r = client.delete(f"/api/people/{create['id']}")
    assert r.status_code == 200
    assert r.json() == {"deleted": 1}
    # Second delete -> 404
    r = client.delete(f"/api/people/{create['id']}")
    assert r.status_code == 404


def test_people_summary_aggregation(client):
    # Rahul: gave 500, then took 200 back (still owes you 300)
    client.post("/api/people/", json={"person_name": "Rahul", "direction": "gave", "amount": 500, "date": "2026-07-01", "reason": "lunch"})
    client.post("/api/people/", json={"person_name": "Rahul", "direction": "took", "amount": 200, "date": "2026-07-05", "reason": "settle partial"})

    # Mom: took 1000 (you owe her 1000)
    client.post("/api/people/", json={"person_name": "Mom", "direction": "took", "amount": 1000, "date": "2026-07-02", "reason": "rent share"})

    # SettledFriend: gave 999 and already marked settled — should be dropped from `people`
    # (no open entries) but kept in `names` for autocomplete.
    client.post(
        "/api/people/",
        json={"person_name": "SettledFriend", "direction": "gave", "amount": 999, "date": "2026-06-01", "settled": True},
    )

    response = client.get("/api/people/summary")
    assert response.status_code == 200
    data = response.json()

    # Global totals derive from per-person net (gave - took), so a person on
    # both sides of the ledger only contributes the net to one side.
    # Rahul net = 500 - 200 = +300 (they owe 300)
    # Mom net   = 0  - 1000 = -1000 (you owe 1000)
    # to_receive = 300, to_give = 1000, net = -700
    assert data["totals"]["to_receive"] == 300.0
    assert data["totals"]["to_give"] == 1000.0
    assert data["totals"]["net"] == -700.0

    # Per-person list: only Rahul and Mom (SettledFriend dropped — all entries settled).
    people = data["people"]
    assert {p["person_name"] for p in people} == {"Rahul", "Mom"}
    # Sort order: largest creditor first → Rahul (+300) before Mom (-1000).
    assert [p["person_name"] for p in people] == ["Rahul", "Mom"]

    rahul = next(p for p in people if p["person_name"] == "Rahul")
    assert rahul["net"] == 300.0
    assert rahul["open_count"] == 2
    assert rahul["gave"] == 500.0
    assert rahul["took"] == 200.0

    mom = next(p for p in people if p["person_name"] == "Mom")
    assert mom["net"] == -1000.0
    assert mom["open_count"] == 1

    # Names list includes SettledFriend for autocomplete, even though they're not in `people`.
    assert "Rahul" in data["names"]
    assert "Mom" in data["names"]
    assert "SettledFriend" in data["names"]

    # Settling the "gave 500" entry means that debt is paid off — it must stop
    # weighing on the net. Only the open "took 200" remains, so Rahul now
    # owes 200 (net flips from +300 to -200). open_count drops to 1.
    gave_entry = next(e for e in rahul["entries"] if e["direction"] == "gave")
    client.put(f"/api/people/{gave_entry['id']}", json={"settled": True})

    data = client.get("/api/people/summary").json()
    rahul = next(p for p in data["people"] if p["person_name"] == "Rahul")
    assert rahul["net"] == -200.0
    assert rahul["open_count"] == 1  # only the "took 200" remains open


def test_people_summary_drops_fully_settled(client):
    client.post(
        "/api/people/",
        json={"person_name": "Ghost", "direction": "gave", "amount": 50, "date": "2026-07-10", "settled": True},
    )
    data = client.get("/api/people/summary").json()
    # All entries settled → dropped from `people`, but kept in `names`.
    assert all(p["person_name"] != "Ghost" for p in data["people"])
    assert "Ghost" in data["names"]
    assert data["totals"]["to_receive"] == 0.0
    assert data["totals"]["to_give"] == 0.0
    assert data["totals"]["net"] == 0.0
